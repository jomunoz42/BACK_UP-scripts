from openpyxl import load_workbook
import re
from collections import defaultdict

wb = load_workbook('buffer.xlsx')
ws = wb.active

# Step 1 - Move column M to A and keep only first word
for row in range(1, ws.max_row + 1):
    value = ws[f"M{row}"].value
    if value is not None:
        # Ensure it's a string before splitting
        first_word = str(value).split()[0] if str(value).strip() != "" else ""
        ws[f"A{row}"].value = first_word
    else:
        ws[f"A{row}"].value = None
    ws[f"M{row}"].value = None  # Clear original col M

# Step 2 - Move column R to B
for row in range(1, ws.max_row + 1):
    value = ws[f"R{row}"].value
    ws[f"B{row}"].value = value
    ws[f"R{row}"].value = None  # Clear original col R

# Step 3 — Keep only columns A and B, delete all others
max_col = ws.max_column
for col in range(max_col, 2, -1):  # start from the end and go down to C
    ws.delete_cols(col)

# Step 4 — Delete rows in which column B contains the given keywords
keywords = ["Villa", "Galé", "Arroios III", "PORTA 12", "Antonio Pedro 2"]

# Loop from bottom to top so deleting rows doesn't mess up indexes
for row in range(ws.max_row, 0, -1):
    cell_value = str(ws[f"B{row}"].value or "")
    if any(keyword in cell_value for keyword in keywords):
        ws.delete_rows(row)

#---------------------------------------------------------------------------------

EDGE_CASES = [
    "501 Dorm 8 male",
    "503 + 505 Dorm 2x8 females",
    "506 Dorm 12 Female",
    "401,301,201"
]

for row in range(2, ws.max_row + 1):
    cell = ws[f"B{row}"]
    val = str(cell.value).strip() if cell.value else ""

    if not val:
        continue

    # NEW: split by commas to process each part independently
    parts = [p.strip() for p in val.split(",")]
    cleaned_parts = []

    for part in parts:
        if "|" in part:
            before_pipe, after_pipe = [x.strip() for x in part.split("|", 1)]

            # If before_pipe contains any edge case keyword → keep only after_pipe
            if any(edge in before_pipe for edge in EDGE_CASES if edge != "(401, 301, 201)"):
                cleaned_parts.append(after_pipe)
            elif "(401, 301, 201)" in before_pipe:
                cleaned_parts.append(after_pipe)
            else:
                # Remove "-" in before_pipe
                before_pipe = before_pipe.replace("-", "")
                cleaned_parts.append(before_pipe + " | " + after_pipe)
        else:
            # No pipe at all — if entire part has an edge case phrase, drop it
            if any(edge in part for edge in EDGE_CASES):
                continue
            else:
                cleaned_parts.append(part)

    # Rejoin cleaned segments
    cell.value = ", ".join(cleaned_parts)

#---------------------------------------------------------------------------

# Regex pattern to keep numbers, (), -, and ,
pattern = re.compile(r"[^0-9(),-]")

# Loop through column B starting from row 2
for row in range(2, ws.max_row + 1):
    cell = ws[f"B{row}"]
    if cell.value is None:
        continue
    if str(cell.value) == "Porteira":
        continue
    # Clean the cell value, keeping allowed characters only
    cleaned = pattern.sub("", str(cell.value))
    cell.value = cleaned

# Remove parentheses if first char != '5'
for row in range(2, ws.max_row + 1):
    cell = ws[f"B{row}"]
    if not cell.value:  # skip empty
        continue

    val = str(cell.value)

    # If first character is NOT 5
    if not val.strip().startswith("5"):
        val = val.replace("(", "").replace(")", "")
        cell.value = val

for row in range(2, ws.max_row + 1):
    cell = ws[f"B{row}"]
    if not cell.value:
        continue
    
    val = str(cell.value).strip()
    
    # Match 503-X or 505-X where X is the bed number
    match = re.match(r"^(503|505)-(\d+)$", val)
    if match:
        dorm = match.group(1)
        bed = match.group(2)
        val = f"{dorm} ({bed})"
    
    cell.value = val

#-------------------------------------------------------------------------

# Final clean-up: Fix leftover 401,301,201 patterns at the end of processing
for row in range(2, ws.max_row + 1):
    cell = ws[f"B{row}"]
    if not cell.value:
        continue

    val = str(cell.value)

    # Replace patterns like 401,301,201301 → 301
    # Also handles multiple occurrences in the same cell
    val = re.sub(r"401,301,201(\d+)", r"\1", val)

    # Remove any accidental double commas and spaces
    val = re.sub(r",\s*,", ",", val)       # fix double commas
    val = re.sub(r"\s*,\s*", ", ", val)    # normalize comma spacing
    val = val.strip(", ").strip()

    cell.value = val

#--------------------------------------------------------------------------

# --------------------------
# Final pretty-grouping pass
# --------------------------
for row in range(2, ws.max_row + 1):
    cell = ws[f"B{row}"]
    if not cell.value:
        continue
    
    val = str(cell.value)
    parts = [p.strip() for p in val.split(",") if p.strip()]

    # Dictionary to hold bed numbers by room
    grouped = {}

    for p in parts:
        # Match patterns like 501(3), 503 (4), 506(12) etc.
        m = re.match(r"^(\d{3})\s*\(?(\d+)\)?$", p)
        if not m:
            continue
        
        room = m.group(1)
        bed = int(m.group(2))
        
        # Special split for 506 (A/B doors)
        if room == "506":
            wing = "A" if 1 <= bed <= 6 else "B"
            key = f"506-{wing}"
        else:
            key = room
        
        grouped.setdefault(key, []).append(bed)

    # Build the pretty result
    pretty_parts = []

    for key in sorted(grouped.keys()):
        beds_sorted = sorted(set(grouped[key]))
        beds_str = ",".join(str(b) for b in beds_sorted)
        
        if key in ["501", "503", "505"]:
            pretty_parts.append(f"{key} ({beds_str})")
        elif key.startswith("506-"):
            pretty_parts.append(f"{key} ({beds_str})")
        else:
            # If it's another room number, just default format
            pretty_parts.append(f"{key} ({beds_str})")

    # Join with comma + space
    cell.value = ", ".join(pretty_parts)

#Well. it worked but the content disappeared


wb.save("buffer.xlsx")

print("Done.")