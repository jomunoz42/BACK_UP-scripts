from openpyxl import load_workbook
from docx import Document

# Load Excel data
wb = load_workbook("buffer.xlsx")
ws = wb.active

names_and_rooms = []
for row in range(2, ws.max_row + 1):
    name = ws[f"E{row}"].value
    room = ws[f"F{row}"].value
    if name and room:
        names_and_rooms.append((name, room))

# Load Word template
doc = Document("word_buffer.docx")
table = doc.tables[0]  # assuming your grid is in the first table

# Fill the table — this example fills across 3 columns per group
index = 0
for row_idx in range(0, len(table.rows), 4):  # every 4 rows is a new "block"
    for col in range(3):  # 3 columns per row group
        if index >= len(names_and_rooms):
            break
        name, room = names_and_rooms[index]
        try:
            # Row 0: Room
            table.cell(row_idx, col).text = f"Room {room}"
            # Row 1: Name
            table.cell(row_idx + 1, col).text = f"Name {name}"
            # Leave code and access untouched or set if you want
        except IndexError:
            pass
        index += 1

# Save new file
doc.save("word_buffer.docx")