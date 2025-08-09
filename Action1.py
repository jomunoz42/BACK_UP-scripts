from openpyxl import load_workbook
from datetime import datetime
import re
from openpyxl.styles import Alignment, Font, Color
from openpyxl.styles.colors import Color as OpenpyxlColor
from openpyxl.styles import PatternFill

wb = load_workbook("buffer.xlsx")
ws = wb.active

# --- Step 0: Copy original column A to column B (using a list buffer) ---
orig_a_values = [ws[f"A{row}"].value for row in range(1, ws.max_row + 1)]
for row, a_value in enumerate(orig_a_values, start=1):
    ws[f"B{row}"] = a_value
# -----------------------------------------------------------------------

# Modify column A
for row in range(1, ws.max_row + 1):
    m_value = ws[f"M{row}"].value or ""
    n_value = ws[f"N{row}"].value or ""
    ws[f"A{row}"] = f"{m_value} {n_value}".strip()

# Step 1: Copy column C to column D
for row in range(1, ws.max_row + 1):
    ws[f"D{row}"] = ws[f"C{row}"].value

# Step 2: Multiply each value in column D by value in column J and then by 4
for row in range(1, ws.max_row + 1):
    d_value = ws[f"D{row}"].value
    j_value = ws[f"J{row}"].value
    d_num = d_value if isinstance(d_value, (int, float)) and d_value is not None else 0
    j_num = j_value if isinstance(j_value, (int, float)) and j_value is not None else 0
    ws[f"D{row}"] = d_num * j_num * 4


today = datetime.now().strftime("%d/%m/%Y")
for row in range(1, ws.max_row + 1):
    ws[f"C{row}"] = today

for row in range(1, ws.max_row + 1):
    m_value = ws[f"M{row}"].value
    if isinstance(m_value, str) and m_value.strip():
        first_word = m_value.strip().split()[0]
    else:
        first_word = ""
    ws[f"E{row}"] = first_word

# Copy column R to column F
for row in range(1, ws.max_row + 1):
    ws[f"F{row}"] = ws[f"R{row}"].value

# Step: Rearrange rows based on conditions in column F

villa_rows = []
arroios_gale_rows = []
other_rows = []

for row in range(1, ws.max_row + 1):
    cell_value = ws[f"F{row}"].value
    cell_str = str(cell_value).strip() if cell_value else ""

    # Check conditions
    if "villa" in cell_str.lower():
        # Rows containing 'villa' anywhere in column F (case-insensitive)
        row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        villa_rows.append(row_values)
    elif cell_str == "Arroios III" or cell_str == "Galé":
        # Rows where column F is exactly 'Arroios III' or 'Galé'
        row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        arroios_gale_rows.append(row_values)
    else:
        # All other rows
        row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        other_rows.append(row_values)

# Combine all rows:
# 1. Other rows first
# 2. Then villa rows
# 3. Then Arroios III and Galé rows
all_rows = other_rows + villa_rows + arroios_gale_rows

# Write them back to worksheet
for row_idx, row_values in enumerate(all_rows, start=1):
    for col_idx, value in enumerate(row_values, start=1):
        ws.cell(row=row_idx, column=col_idx).value = value

# Clear any leftover rows if the sheet originally had more rows
new_max_row = len(all_rows)
old_max_row = ws.max_row
if old_max_row > new_max_row:
    for clear_row in range(new_max_row + 1, old_max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=clear_row, column=col).value = None

# Center alignment style
center_align = Alignment(horizontal='center')

# Bold font with dark blue color (Azul Escuro 1 approx)
bold_blue_font = Font(bold=True, color="264478") 

# Apply center alignment to columns A to F
for row in range(1, ws.max_row + 1):
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f"{col_letter}{row}"].alignment = center_align

# Apply bold font and blue color to column D only
for row in range(1, ws.max_row + 1):
    ws[f"D{row}"].font = bold_blue_font


# Define fills for the backgrounds
# Much lighter gray for Villa rows
fill_villa = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Keep the current color for Arroios/Galé
fill_gale_arroios = PatternFill(start_color="D1D1D1", end_color="D1D1D1", fill_type="solid")

max_col = ws.max_column
max_row = ws.max_row

for row in range(1, max_row + 1):
    cell_value = ws[f"F{row}"].value
    cell_str = str(cell_value).strip() if cell_value else ""

    if "villa" in cell_str.lower():
        # Apply fill to entire row for Villa rows
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).fill = fill_villa

    elif cell_str == "Arroios III" or cell_str == "Galé":
        # Apply fill to entire row for Arroios III or Galé
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).fill = fill_gale_arroios

# Step 1: Delete column E values until you find a colored cell (not white)
row = 1
while True:
    cell_e = ws[f"E{row}"]
    fill = cell_e.fill
    color = fill.fgColor.rgb if fill.patternType else None
    if color and color != "FFFFFFFF":  # Found first colored cell
        break
    cell_e.value = None
    row += 1

# Step 2: Copy from D to E while E has color, and apply desired color + bold
custom_blue = "FF264478"  # ARGB format for openpyxl (FF for full opacity)
while True:
    cell_e = ws[f"E{row}"]
    fill = cell_e.fill
    color = fill.fgColor.rgb if fill.patternType else None
    if not color or color == "FFFFFFFF":  # Stop at first white cell in E
        break
    left_cell = ws[f"D{row}"]
    
    # Copy value
    cell_e.value = left_cell.value
    # Apply font style (bold + blue)
    cell_e.font = Font(name=left_cell.font.name or "Calibri",
                       size=left_cell.font.size or 11,
                       bold=True,
                       color=custom_blue)
    row += 1

# Step 3: Erase all contents in columns F, G, and H
for col in ['F', 'G', 'H']:
    for r in range(1, ws.max_row + 1):
        ws[f"{col}{r}"].value = None


wb.save("buffer.xlsx")

print("Done.")