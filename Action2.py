from openpyxl import load_workbook
from datetime import datetime
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
# import win32com.client


wb = load_workbook("buffer.xlsx")
ws = wb.active

max_row = ws.max_row

# Loop through each row and move data
for row in range(1, max_row + 1):

    ws[f"A{row}"] = ws[f"M{row}"].value

    ws[f"C{row}"] = ws[f"K{row}"].value

    ws[f"B{row}"] = ws[f"J{row}"].value

today_str = datetime.today().strftime('%Y-%m-%d')

for row in range(ws.max_row, 0, -1):
    cell_value = ws[f"D{row}"].value
    if isinstance(cell_value, datetime):
        cell_str = cell_value.strftime('%Y-%m-%d')
    else:
        cell_str = str(cell_value)
    if cell_str == today_str:
        ws.delete_rows(row)

for row in range(1, max_row + 1):
    ws[f"D{row}"] = ws[f"R{row}"].value

for row in range(1, max_row + 1):
    ws[f"E{row}"] = ws[f"AB{row}"].value

max_col = ws.max_column
start_col = column_index_from_string('F')
ws.delete_cols(start_col, max_col - start_col + 1)

import re

to_delete = ["Villa", "PORTA 10", "PORTA 12", "Antonio Pedro", "Arroios", "Galé"]

for row in range(ws.max_row, 1, -1):  # start from 2 to preserve header (row 1)
    cell = ws[f"D{row}"].value
    if cell is None:
        continue
    if any(term in str(cell) for term in to_delete):
        ws.delete_rows(row)

# Step 2: Trim special dorm cells
for row in range(2, ws.max_row + 1):  # again skip row 1 (immune)
    cell = ws[f"D{row}"].value
    if cell is None:
        continue
    cell_str = str(cell)

    # Handle exact special cases with '|'
    if '|' in cell_str and "503 + 505 Dorm" in cell_str:
        parts = cell_str.split('|')
        if len(parts) > 1:
            after_pipe = parts[1].strip()
            if '-' in after_pipe:
                new_val = after_pipe.split('-')[0]
                ws[f"D{row}"] = new_val
            else:
                ws[f"D{row}"] = after_pipe
        continue

    # Handle cases starting with dorm names
    if cell_str.startswith("501 Dorm 8 male"):
        ws[f"D{row}"] = "501"
    elif cell_str.startswith("506 Dorm 12 Female"):
        ws[f"D{row}"] = "506"

# Step 3: Extract only numbers (except row 1)
import re

for row in range(2, ws.max_row + 1):
    cell = ws[f"D{row}"].value
    if cell is None:
        continue
    cell_str = str(cell)

    # Only keep numeric content if not already a known good code
    if not cell_str in ["501", "503", "505", "506"]:
        numbers = re.findall(r'\d+', cell_str)
        if numbers:
            ws[f"D{row}"] = numbers[-1]  # keep the last number
        else:
            ws[f"D{row}"] = ""

#delete rows "sem PA"
for row in range(ws.max_row, 1, -1):  # start from bottom, skip header row 1
    cell = ws[f"E{row}"].value
    if cell is None:
        continue
    if "sem PA" in str(cell):
        ws.delete_rows(row)

for row in range(1, ws.max_row + 1):
    ws[f"E{row}"] = None

for row in range(2, ws.max_row + 1):
    cell = ws[f"A{row}"].value
    if cell is None:
        continue
    first_word = str(cell).split()[0]
    ws[f"A{row}"] = first_word

# Step: Sort rows by numeric value in column D (excluding header)

# Get all data rows (excluding header)
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(list(row))

# Sort the data by column D (index 3), converting to int safely
data.sort(key=lambda row: int(str(row[3])) if row[3] is not None and str(row[3]).isdigit() else float('inf'))

# Overwrite the sheet starting from row 2
for i, row_data in enumerate(data, start=2):
    for j, value in enumerate(row_data, start=1):
        ws.cell(row=i, column=j, value=value)


# Step 1: Find the last non-empty row in column A
last_row = ws.max_row
for row in range(ws.max_row, 0, -1):
    if ws[f"A{row}"].value not in [None, ""]:
        last_data_row = row
        break

# Step 2: Target row (2 rows below last content)
target_row = last_data_row + 2

# Step 3: Write "TOTAL" in A
ws[f"A{target_row}"] = "TOTAL"
ws[f"A{target_row}"].font = Font(bold=True)
ws[f"A{target_row}"].alignment = Alignment(horizontal="center")

# Step 4: Write sum formula in B
ws[f"B{target_row}"] = f"=SUM(B2:B{last_data_row}) + SUM(C2:C{last_data_row})"
ws[f"B{target_row}"].font = Font(bold=True)
ws[f"B{target_row}"].alignment = Alignment(horizontal="center")

for col in range(1, 5):  # Columns A to D
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

for row in range(2, ws.max_row + 1):
    ws[f"D{row}"].alignment = Alignment(horizontal="right")

ws.print_options.gridLines = True

ws.column_dimensions['A'].width = 21
ws.column_dimensions['B'].width = 13
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 35


# wb = load_workbook("buffer.xlsx")

wb.save("buffer.xlsx")

# excel = win32com.client.Dispatch("Excel.Application")
# excel.Visible = False
# wb_path = r"C:\Users\35196\Documents\My Games\booking_back_up\Scripts.py\hello.xlsx"
# workbook = excel.Workbooks.Open(wb_path)
# workbook.Worksheets(1).PrintOut(ActivePrinter="Brother HL-L2350DW series Printer")
# workbook.Close(SaveChanges=False)
# print("Current active printer is:", excel.ActivePrinter)
# excel.Quit()


print("Done.")