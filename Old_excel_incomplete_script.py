from openpyxl import load_workbook
from datetime import datetime
import re

wb = load_workbook("hello.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    row[1].value = row[0].value

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    val_m = row[12].value
    val_n = row[13].value
    val_m = str(val_m) if val_m is not None else ""
    val_n = str(val_n) if val_n is not None else ""
    row[0].value = val_m + " " + val_n

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    val_c = row[2].value  # Column C (index 2)
    val_j = row[9].value  # Column J (index 9)
    val_c = val_c if val_c is not None else 0
    val_j = val_j if val_j is not None else 0
    try:
        row[3].value = val_c * val_j * 4  # Column D (index 3)
    except TypeError:
        row[3].value = "ERROR"

today = datetime.today().strftime("%d/%m/%Y")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row[2].value = today  # Column C (index 2)


wb.save("hello.xlsx")

print("Done.")
